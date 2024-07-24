import os
from datetime import date
from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox

import openpyxl
from PIL import ImageTk, Image
from openpyxl import Workbook
import pathlib

BACKGROUND_COLOR = "#06283D"
FRAME_BG_COLOR = "#EDEDED"
FRAME_FG_COLOR = "#06283D"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=BACKGROUND_COLOR)

file_path = pathlib.Path("Student_data.xlsx")
if not file_path.exists():
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        "Registration No.", "Name", "Class", "Gender", "DOB",
        "Date Of Registration", "Religion", "Skill",
        "Father Name", "Mother Name", "Father's Occupation", "Mother's Occupation"
    ]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    workbook.save(file_path)



########################Search#####################
def search():

    text = Search.get() #taking input from entry box

    Clear() #to clear all the data already available in entry box
    save_button.config(state="disable") #after clicking on search, save button will disable so that no one can click on it

    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            ## print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number=str(name)[15:-1]

            # print(reg_no_position)
            # print(reg_number)

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid", "Invalid registration number!!!")

    #reg_no_position showing like A2, A3, A4 ...
    #but reg_number just showing number after A2 like 2, 3 ...

    x1=sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value

    # print(x1)
    # print(x2)
    # print(x3)
    # print(x4)
    # print(x5)
    # print(x6)
    # print(x7)
    # print(x8)
    # print(x9)
    # print(x10)
    # print(x11)
    # print(x2)

    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)

    if x4=="Female":
        R2.select()
    else:
        R1.select()

    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    Father_Occupation.set(x11)
    Mother_Occupation.set(x12)

    img1 = (Image.open("media/"+str(x1)+".jpg")) #show images in media file, as their reg_number
    resized_image1=img1.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image1)
    lbl.config(image=photo2)
    lbl.image=photo2


######################Update#######################
def Update():
    # print("works")
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()






# Gender
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
        print(gender)
    else:
        gender = "Female"
        print(gender)


# Exit command
def Exit():
    root.destroy()


# Show Image
def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select image file", filetypes=(("JPG File", "*.jpg"),
                                                                               ("PNG File", "*.png"),
                                                                               ("All files", "*.*")))
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2


# Registration No.
def registration_no():
    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)

    except:
        Registration.set("1")


# Clear
def Clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set('Select Class')

    registration_no()

    save_button.config(state='normal')

    img1 = Image.open("/Users/ivan/Desktop/coding/programming/petprojects/Student_RegistrationSystem/media/profile.png")
    resized_image = img1.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

    img = ""


# Save
def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("error", "Select Gender!")
        return

    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    if N1 == "" or C1 == "Select Class" or D2 == "" or Re1 == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("error", "Few Data is missing!")
        return

    else:
        file = openpyxl.load_workbook("Student_data.xlsx")
        sheet = file.active
        new_row = sheet.max_row + 1
        sheet.cell(column=1, row=new_row, value=R1)
        sheet.cell(column=2, row=new_row, value=N1)
        sheet.cell(column=3, row=new_row, value=C1)
        sheet.cell(column=4, row=new_row, value=G1)
        sheet.cell(column=5, row=new_row, value=D2)
        sheet.cell(column=6, row=new_row, value=D1)
        sheet.cell(column=7, row=new_row, value=Re1)
        sheet.cell(column=8, row=new_row, value=S1)
        sheet.cell(column=9, row=new_row, value=fathername)
        sheet.cell(column=10, row=new_row, value=mothername)
        sheet.cell(column=11, row=new_row, value=F1)
        sheet.cell(column=12, row=new_row, value=M1)
        file.save('Student_data.xlsx')

        try:
            img.save(f"media/{R1}.jpg")
        except:
            messagebox.showinfo("info", "Profile picture is not available!")

        messagebox.showinfo("info", "Data entered successfully!")

        Clear()  # clear entry box and image section
        registration_no()  # recheck registration no. and reissue new no.


# Top frames
Label(root, text="Email: nazaruk649@ukr.net", width=10, height=3, bg='#f0687c', anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg='#c36464', fg="#fff", font='arial 20 bold').pack(side=TOP, fill=X)

# Search box to update
Search = StringVar()
Entry(root, textvariable=Search, width=17, bd=2, font='arial 20').place(x=820, y=65)
original_image = Image.open("/Users/ivan/Desktop/coding/programming/petprojects/Student_RegistrationSystem/media/search.png")
resized_image = original_image.resize((20, 20))
imageicon3 = ImageTk.PhotoImage(resized_image)

Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=100, height=30, bg="#ADD8E6", font="arial 13 bold", command=search)
Srch.place(x=1060, y=65)

# Update button
original_image2 = Image.open("/Users/ivan/Desktop/coding/programming/petprojects/Student_RegistrationSystem/media/update.png")
resized_image = original_image2.resize((40, 40))
imageicon4 = ImageTk.PhotoImage(resized_image)

Update_button = Button(root, image=imageicon4, bg="#c36464", borderwidth=0, command=Update)
Update_button.place(x=110, y=60)

# Registration and Date
Label(root, text="Registration No:", font="arial 13", fg="white", bg=BACKGROUND_COLOR).place(x=30, y=150)
Label(root, text="Date:", font="arial 13", fg="white", bg=BACKGROUND_COLOR).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font='arial 10')
reg_entry.place(x=160, y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)

Date.set(d1)

# Student details
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=BACKGROUND_COLOR, fg='white', height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(root, text='Full Name:', font='arial 13', bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR).place(x=40, y=250)
Label(root, text='Date of Birth:', font='arial 13', bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR).place(x=40, y=300)
Label(root, text='Gender:', font='arial 13', bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR).place(x=40, y=350)

Label(root, text='Class:', font='arial 13', bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR).place(x=500, y=250)
Label(root, text='Religion:', font='arial 13', bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR).place(x=500, y=300)
Label(root, text='Skill:', font='arial 13', bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR).place(x=500, y=350)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font='arial 10')
name_entry.place(x=160, y=30)

DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=20, font='arial 10')
dob_entry.place(x=160, y=80)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR, command=selection)
R1.place(x=150, y=133)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR, command=selection)
R2.place(x=230, y=133)

Religion = StringVar()
religion_entry = Entry(obj, textvariable=Religion, width=20, font='arial 10')
religion_entry.place(x=630, y=80)

Skill = StringVar()
skill_entry = Entry(obj, textvariable=Skill, width=20, font='arial 10')
skill_entry.place(x=630, y=135)

Class = Combobox(obj, values=["ICT", "Math", "Science", "English", "History", "Geography", "Art", "Physical Education"], font="Roboto 10", width=17, state='r')
Class.place(x=630, y=30)
Class.set("Select Class")

# Parent details
parent_frame = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=BACKGROUND_COLOR, fg='white', height=220, relief=GROOVE)
parent_frame.place(x=30, y=470)

Label(parent_frame, text="Father's Name:", font="arial 13", bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR).place(x=30, y=50)
Label(parent_frame, text="Occupation:", font="arial 13", bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR).place(x=30, y=100)

F_Name = StringVar()
f_name_entry = Entry(parent_frame, textvariable=F_Name, width=20, font='arial 10')
f_name_entry.place(x=150, y=50)

Father_Occupation = StringVar()
FO_entry = Entry(parent_frame, textvariable=Father_Occupation, width=20, font='arial 10')
FO_entry.place(x=150, y=100)

Label(parent_frame, text="Mother's Name:", font="arial 13", bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR).place(x=500, y=50)
Label(parent_frame, text="Occupation:", font="arial 13", bg=BACKGROUND_COLOR, fg=FRAME_BG_COLOR).place(x=500, y=100)

M_Name = StringVar()
m_name_entry = Entry(parent_frame, textvariable=M_Name, width=20, font='arial 10')
m_name_entry.place(x=620, y=50)

Mother_Occupation = StringVar()
MO_entry = Entry(parent_frame, textvariable=Mother_Occupation, width=20, font='arial 10')
MO_entry.place(x=620, y=100)

# Image Frame
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

original_image = Image.open("/Users/ivan/Desktop/coding/programming/petprojects/Student_RegistrationSystem/media/profile.png")
resized_image = original_image.resize((200, 200))
img_profile = ImageTk.PhotoImage(resized_image)
lbl = Label(f, bg="black", image=img_profile)
lbl.place(x=0, y=0)

# Button
upload_button = Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage)
upload_button.place(x=1000, y=370)

save_button = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
save_button.place(x=1000, y=450)

reset_button = Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear)
reset_button.place(x=1000, y=530)

exit_button = Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit)
exit_button.place(x=1000, y=610)

root.mainloop()
